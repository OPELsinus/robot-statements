import csv
import logging
import os
import re
import time
import shutil
import traceback
from contextlib import suppress
from pathlib import Path

import keyboard
import numpy as np
import pandas as pd
import datetime

import pyperclip
import xlwings as xw
from xlwings.constants import VAlign, HAlign

from openpyxl import load_workbook
from core import Odines

import tools
from tools import hold_session, send_message_by_smtp, send_message_to_orc, update_credentials, send_message_to_tg
from config import smtp_host, smtp_author, chat_id, download_path, working_path, SEDLogin, SEDPass, save_xlsx_path, owa_username, owa_password, logger_name, save_xlsx_path_qlik, tg_token, machine_ip, halyk_extract_path, halyk_extract_path_2023, process_day
from rpamini import Web

cols = ['N', 'Согласован', 'Дата выписки', 'Дата планируемой оплаты', 'Заявка на оплату',
        'Вид операции', 'Организация', 'Контрагент', 'БИН / ИИН', 'Статья затрат', 'Код БДДС', 'Документ основание', 'Валюта документа', 'Сумма документа',
        'Заявка на расходование', 'Платежное поручение']

MONTHS = [
    'Январь',
    'Февраль',
    'Март',
    'Апрель',
    'Май',
    'Июнь',
    'Июль',
    'Август',
    'Сентябрь',
    'Октябрь',
    'Ноябрь',
    'Декабрь'
]


class Registry(Web):

    def __init__(self):
        super(Registry, self).__init__()

    def load(self):
        selector_ = '//div[@id="thinking" and contains(@style, "block")]'
        self.wait_element(selector_, timeout=2)

        selector_ = '//div[@id="thinking" and contains(@style, "none")]'
        self.wait_element(selector_)

    def auth(self):
        self.get('http://172.16.10.86/user/login')

        self.find_element('//input[@id="login"]').type_keys(SEDLogin)
        self.find_element('//input[@id="password"]').type_keys(SEDPass)
        self.find_element('//input[@id="submit"]').click()

        self.wait_element('//span[@class="user_shortinfo_infoname"]')
        return self


def search_by_date(yest):
    # print('Started searching by date')
    # send_message_to_orc('https://rpa.magnum.kz/tg', chat_id, 'Начат фильтр по дате')
    web = Registry()
    web.run()
    web.auth()
    web.load()

    web.find_element('//*[@id="header_menu"]/li[7]/a').click()
    time.sleep(0.2)

    # Кнопка "Документы -> Финансовые"
    web.find_element('//a[contains(text(), "Финансовые")]').click()
    time.sleep(0.1)

    # Кнопка "Документы -> Финансовые -> Реестр на оплату"
    web.find_element('//a[contains(text(), "Реестр на оплату")]').click()

    web.load()

    print()
    web.find_element("//div[text()='Оплата']").click()

    # Фильтр дял поиска по дате и виду реестра
    web.find_element('//*[@id="content_top_search_bar"]/ul[1]/li[3]/span/i[1]').click()
    print()
    time.sleep(0.2)
    try:
        web.find_element('//*[@id="extended_search_container_folder"]/div/div[2]/div/div[1]/div/input').click()
    except:
        web.find_element('//*[@id="content_top_search_bar"]/ul[1]/li[3]/span/i[1]').click()
        time.sleep(0.2)
        web.find_element('//*[@id="extended_search_container_folder"]/div/div[2]/div/div[1]/div/input').click()
    print()
    time.sleep(0.1)
    # print(yest)
    web.find_element('//*[@id="extended_search_container_folder"]/div/div[2]/div/div[1]/div/ul/li[6]').click()
    time.sleep(0.2)
    print()
    web.find_element('//*[@id="filterPeriodFrom_date_widget"]').click()
    web.find_element('//*[@id="filterPeriodFrom_date_widget"]').set_attr(yest.replace('.', ''))

    web.find_element('//*[@id="filterPeriodTo_date_widget"]').click()
    web.find_element('//*[@id="filterPeriodTo_date_widget"]').set_attr(yest.replace('.', ''))

    print()
    time.sleep(0.13)

    # OK Button
    web.find_element('/html/body/div[9]/div[3]/div/button[2]').click()

    time.sleep(0.1)

    print()
    for tries in range(3):
        try:
            web.find_element('//*[@id="extended_search_container_folder"]/div/div[2]/div/div[5]/input[4]').click()
            web.find_element('//*[@id="extended_search_container_folder"]/div/div[2]/div/div[5]/input[4]').set_attr('Безналичный')
            break
        except Exception as error:
            print(f"ERROR OCCURED HERE: {error}")
            # time.sleep(1000)
        time.sleep(0.1)

    print('here1')
    web.find_element('//*[@id="extended_search_container_folder"]/div/div[2]/div/div[6]/input[4]').click()
    time.sleep(0.1)

    web.find_element('//*[@id="extended_search_container_folder"]//button[text()="Найти"]').click()
    time.sleep(0.1)

    print('here2')
    return web


def documentolog(web, yesterday):

    try:
        print('enter1')
        if web.find_element('//*[@id="node_meta_total_rows"]').get_attr('text') == '0':
            web.quit()
            print('quitted')
            return [None, None]
        print('enter2')
        time.sleep(1.5)
        order = web.find_element('//*[contains(@id, "grid_col_f")]/span/a[2]').get_attr('class')

        print('enter2.1')
        # Сортировка реестров по убыванию даты
        if 'desc' in order and 'current' not in order:
            print('enter2.2')
            web.find_element('//*[@id="grid_col_f_4121eee"]/span').click()

            print('enter2.3')
            time.sleep(1.5)
        time.sleep(0.1)
        texts = []
        links = []
        times = []
        link = []
        rows = []
        start = 0
        end = 9
        end_ = 9
        print('enter3')
        cells = web.find_elements('//*[contains(@id, "grid_cell")]/a')

        print('enter3.00')
        for ind, cell in enumerate(cells):
            try:
                if ind % end == 0:
                    links.append(cell.get_attr('href'))
                texts.append(cell.get_attr('text'))

            except:
                print('eroro')
                send_message_to_orc('https://rpa.magnum.kz/tg', chat_id, 'Сверка выписок\nОШИБКА1')

        print('enter3.1')
        while end <= len(texts):
            # print(f"Appending {texts[start:end]} | {start} | {end}")
            rows.append(texts[start:end])
            start, end = end, end + end_

        today = datetime.datetime.strptime(datetime.date.today().strftime('%d.%m.%Y'), '%d.%m.%Y').date()
        print('enter4')
        if len(yesterday) == 0:
            yesterd_reestr_date = ''
        else:
            yesterd_reestr_date = yesterday

        df2 = pd.DataFrame()

        print('enter4.1')
        # print(rows)
        # send_message_to_orc('https://rpa.magnum.kz/tg', chat_id, f'Начат отбор реестров. Всего {len(links)} реестра(-ов)')
        for ind, row in enumerate(rows):
            # print(row)
            row_date = datetime.datetime.strptime(row[0], '%d.%m.%Y').date()
            # print('enter4.2', row_date)
            if row_date < today:
                logger.info(f'Checking')
                if len(yesterd_reestr_date) == 0:
                    yesterd_reestr_date = row[0]

                if len(yesterd_reestr_date) != 0 and row[0] == yesterd_reestr_date and 'Безналичный' in row and row_date == datetime.datetime.strptime(row[0], '%d.%m.%Y').date():
                    start_time1 = datetime.datetime.now().strftime('%H:%M:%S')

                    web.get(links[ind])
                    web.load()

                    link.append(links[ind])
                    # logger.info(f'Started reestr: {links[ind]}')
                    df1 = get_data_from_reestr(web)
                    # logger.info(f'Ended reestr: {links[ind]}')
                    df2 = pd.concat([df2, df1])
                    # logger.info(f'Concatenated')
                    end_time = datetime.datetime.now().strftime('%H:%M:%S')
                    times.append([start_time1, end_time])
                    # print(row, links[ind])

        logger.info(f'Went forward')

        # ----------------------------------------------------------------------------------
        # Выполнение кода до страницы 11 ТЗ
        # ----------------------------------------------------------------------------------

        # Кнопка "Справочники"
        web.find_element('//*[@id="header_menu"]/li[6]/a').click()
        time.sleep(0.7)

        # Кнопка "Справочники -> Системные"
        web.find_element('//*[@id="header_menu"]/li[6]/div/ul/li[4]/a').click()
        time.sleep(0.7)

        # Кнопка "Справочники -> Системные -> Список файлов"
        web.find_element('//*[@id="header_menu"]/li[6]/div/ul/li[4]/div/ul/li[6]/a').click()
        time.sleep(1)

        year = str(yesterday).split('.')[2]

        df1 = pd.DataFrame()

        for index in range(100):

            title = web.find_element(f'//*[@id="grid_row_{index}"]/td[2]/a').get_attr('text')

            # if 'Факт' in title and 'оплат' in title and year in title:
            if 'Факт' in title and 'оплат' in title and year in title:

                web.find_element(f'//*[@id="grid_row_{index}"]/td[2]/a').click()
                web.find_element('//*[contains(@id, "fileview")]').click()

                filename = None
                found = False

                for wait in range(60):
                    for file_ in os.listdir(download_path):
                        if 'факт' in file_.lower() and 'оплат' in file_.lower() and 'crdownload' not in file_.lower():
                            filename = file_
                            found = True
                            break
                    if found:
                        break
                    else:
                        time.sleep(1)

                # print(filename)
                df1 = fact_oplat_to_reestr(filename, yesterd_reestr_date)

                logger.info('Deleting')
                Path.unlink(Path(os.path.join(download_path, filename)))
                logger.info('Deleted')

                break
                # send_message_to_orc('https://rpa.magnum.kz/tg', chat_id, 'Сверка выписок\nОШИБКА2')
        # except:
        #     break

        df2 = pd.concat([df2, df1])

        # for times1 in times:
        #     print(times1)
        # print(yesterd_reestr_date)

        web.quit()
        return [df2, yesterd_reestr_date]

    except Exception as error1:
        logger.exception(f'Error Occured on Documentolog: {error1}')
        traceback.print_exc()


def get_data_from_reestr(web):
    # print('Started reestr')

    reestr_title = web.find_element('//*[@id="reference-view"]/table/tbody/tr[3]/td[2]').get_attr('text')

    if 'таткрафт' in str(reestr_title).lower():
        print('skipped tatcraft')
        return pd.DataFrame()

    provider_name = []
    provider_bin_iin = []
    statement_in_dds = []
    payment_currency = []
    amount_to_pay = []
    payment_date = []

    cells1 = web.find_elements('//tr[@class="group-name trow"]/td[5]')
    cells2 = web.find_elements('//tr[@class="group-name trow"]/td[4]')
    cells3 = web.find_elements('//tr[@class="group-name trow"]/td[14]')
    logger.info('Found 3 cells')
    hold_session()
    cells4 = web.find_elements('//tr[@class="group-name trow"]/td[9]')
    cells5 = web.find_elements('//tr[@class="group-name trow"]/td[8]')
    cells6 = web.find_elements('//tr[@class="group-name trow"]/td[7]')
    hold_session()
    logger.info('Found 6 cells')

    for id, cell in enumerate(cells1):
        try:
            text = cell.get_attr('text').strip()
            # if len(cells2[id + 1].get_attr('text').strip()) != 0:
            #     provider_name.append(text) if len(text) != 0 else provider_name.append(' ')
            # else:
            provider_name.append(text) if len(text) != 0 else provider_name.append('')
            # print(text)
        except Exception as error1:
            logger.warning(f'Error Occured on Reestr: {error1}')
            logger.info(f'Error Occured on Reestr: {error1}')
            traceback.print_exc()
    print('~~1~~')
    for cell in cells2:
        try:
            text = cell.get_attr('text').strip()
            provider_bin_iin.append(text) if len(text) != 0 else provider_bin_iin.append('')
        except Exception as error1:
            logger.warning(f'Error Occured on Reestr2: {error1}')
            logger.info(f'Error Occured on Reestr2: {error1}')
            traceback.print_exc()
    for cell in cells3:
        try:
            text = cell.get_attr('text').strip()
            statement_in_dds.append(text) if len(text) != 0 else statement_in_dds.append('')

        except Exception as error1:
            logger.warning(f'Error Occured on Reestr3: {error1}')
            logger.info(f'Error Occured on Reestr3: {error1}')
            traceback.print_exc()

    logger.info('Appended 3 cells')
    hold_session()
    for cell in cells4:
        try:
            text = cell.get_attr('text').strip()
            payment_currency.append(text) if len(text) != 0 else payment_currency.append('')

        except Exception as error1:
            logger.warning(f'Error Occured on Reestr4: {error1}')
            logger.info(f'Error Occured on Reestr4: {error1}')
            traceback.print_exc()
    for cell in cells5:
        try:
            text = cell.get_attr('text').strip()
            amount_to_pay.append(text) if len(text) != 0 else amount_to_pay.append('')

        except Exception as error1:
            logger.warning(f'Error Occured on Reestr5: {error1}')
            logger.info(f'Error Occured on Reestr5: {error1}')
            traceback.print_exc()
    for cell in cells6:
        try:
            text = cell.get_attr('text').strip()
            payment_date.append(text) if len(text) != 0 else payment_date.append('')

        except Exception as error1:
            logger.warning(f'Error Occured on Reestr6: {error1}')
            logger.info(f'Error Occured on Reestr6: {error1}')
            traceback.print_exc()
    logger.info('Appended 6 cells')

    if 'го' in reestr_title.lower() and 'доп' not in reestr_title.lower():
        reestr_title = 'Реестр (ГО)'

    elif 'го' in reestr_title.lower() and 'доп' in reestr_title.lower():
        reestr_title = 'Дополнительный реестр (ГО)'

    elif 'филиал' in reestr_title.lower() and 'доп' not in reestr_title.lower():
        reestr_title = 'Реестр (Филиалы)'

    elif 'филиал' in reestr_title.lower() and 'доп' in reestr_title.lower():
        reestr_title = 'Дополнительный реестр (Филиалы)'

    elif 'инвест' in reestr_title.lower() and 'доп' not in reestr_title.lower():
        reestr_title = 'Реестр (Инвест)'

    elif 'инвест' in reestr_title.lower() and 'доп' in reestr_title.lower():
        reestr_title = 'Дополнительный реестр (Инвест)'

    elif 'магнум астана' in reestr_title.lower() and 'доп' not in reestr_title.lower():
        reestr_title = 'Магнум Астана'

    elif 'магнум астана' in reestr_title.lower() and 'доп' in reestr_title.lower():
        reestr_title = 'Дополнительный реестр (Магнум Астана)'

    elif 'реестр пф' in reestr_title.lower() and 'доп' not in reestr_title.lower():
        reestr_title = 'Реестр ПФ'

    elif 'реестр 1' in reestr_title.lower() and 'доп' not in reestr_title.lower():
        reestr_title = 'Реестр 1С'

    statement_check = []
    print('STATEMENTS IN DDS:', statement_in_dds)
    print()
    for ind, string in enumerate(statement_in_dds):

        try:
            statement_in_dds[ind] = string.split(';')[0]
        except:
            # statement_in_dds[ind] = ''
            statement_check.append('')
            continue
        try:
            statement_check.append(string.split(';')[1].strip())
        except:
            statement_check.append(string)

    for ind in range(len(payment_date)):
        try:
            payment_date[ind] = payment_date[ind].strip()
            payment_date[ind] = payment_date[ind][:6] + payment_date[ind][-2:]
        except:
            continue

    try:
        amount_to_pay = [s.replace(' ', '') for s in amount_to_pay]
        amount_to_pay = np.asarray(amount_to_pay).astype(float)
    except:
        ...
    # for ind, j in enumerate(provider_name):
    #     print(ind, j)
    print(len(provider_name), len(provider_bin_iin), len(reestr_title), len(reestr_title), len(statement_in_dds), len(payment_currency), len(amount_to_pay), len(payment_date), len(statement_in_dds), len(statement_check))
    print('--------=====----==')
    # print((provider_name), (provider_bin_iin), (reestr_title), reestr_title, (statement_in_dds), (payment_currency), (amount_to_pay), (payment_date), (statement_in_dds), sep='\n')
    if len(reestr_title) != len(provider_name):
        reestr_title = [reestr_title] * len(provider_name)
    print('-==-=-=-=-=--=--=-=-=-=')
    print(len(provider_name), len(provider_bin_iin), len(reestr_title), len(reestr_title), len(statement_in_dds), len(payment_currency), len(amount_to_pay), len(payment_date), len(statement_in_dds), len(statement_check))
    print('--------=====----==')
    # print((provider_name), (provider_bin_iin), (reestr_title), reestr_title, (statement_in_dds), (payment_currency), (amount_to_pay), (payment_date), (statement_in_dds), sep='\n')

    df3 = pd.DataFrame({
        'Поставщик': provider_name,
        'БИН / ИИН получателя': provider_bin_iin,
        'Реестр': reestr_title,
        'Статья в ДДС': statement_in_dds,
        'Валюта платежа': payment_currency,
        'Сумма к оплате': amount_to_pay,
        'Сумма к оплате KZT': amount_to_pay,
        'Курс': [1] * len(provider_name),
        'Дата оплаты': payment_date,
        'Skip': [''] * len(provider_name),
        'Проверка статьи': statement_check,
        'Название статьи': statement_in_dds
    })
    # print(df3)
    hold_session()
    # logger.info(f'DF Length: {len(df3)}')
    return df3

    # except:
    #     send_message_to_orc('https://rpa.magnum.kz/tg', chat_id, 'Сверка выписок\nОШИБКА3')


def fact_oplat_to_reestr(filename, yesterdays_reestr_date):
    print('STARTED FACT OPLAT')
    print(filename, yesterdays_reestr_date)
    # print('Started fact oplat to reestr')
    hold_session()
    time.sleep(3)
    df_ = pd.read_excel(os.path.join(download_path, filename))
    print(df_)
    # print(os.path.join(download_path, filename))
    df_['Дата'] = pd.to_datetime(df_['Дата'], format='%d.%m.%Y')
    df_['Дата'] = df_['Дата'].dt.strftime('%d.%m.%y')

    yesterdays_reestr_date = yesterdays_reestr_date[:6] + yesterdays_reestr_date[-2:]
    df_ = df_[(df_['Оплата'] == 'Б/Н') & (df_['Дата'] == yesterdays_reestr_date)]

    try:
        df_['Сумма на оплату'] = df_['Сумма на оплату'].apply(lambda x: x.replace(' ', ''))
    except:
        ...

    df_['Сумма на оплату'] = df_['Сумма на оплату'].astype(float)

    return pd.DataFrame({'Поставщик': df_['Наименование поставщика'], 'БИН / ИИН получателя': '', 'Реестр': '', 'Статья в ДДС': '', 'Валюта платежа': 'KZT',
                         'Сумма к оплате': df_['Сумма на оплату'], 'Сумма к оплате KZT': df_['Сумма на оплату'], 'Курс': 1, 'Дата оплаты': df_['Дата'], 'Skip': '', 'Проверка статьи': df_['Наименование поставщика'], 'Название статьи': ''})


def get_first_statement(weekends):

    hold_session()
    # print('Started getting first statement')
    df1_ = pd.DataFrame()
    print(weekends)

    for ind, day in enumerate(weekends[::-1]):

        month = int(day.split('.')[1])
        year = str(day.split('.')[2])

        # for folder in os.listdir(halyk_extract_path):
        #     print('#1', month, folder, MONTHS[month - 1], year)
        #     if str(month) in folder and MONTHS[month - 1] in folder and str(year) in folder:
        #         for subfolders in os.listdir(os.path.join(halyk_extract_path, folder)):
        #
        #             subfolder = os.path.join(os.path.join(halyk_extract_path, folder), subfolders)
        #             try:
        #                 for files in os.listdir(subfolder):
        #
        #                     file = os.path.join(os.path.join(os.path.join(halyk_extract_path, folder), subfolders), files)
        #
        #                     print('#2', day, files.lower(), os.path.getsize(file) / 1024)
        #                     if day in files and 'kzt народный' in files.lower() and os.path.getsize(file) / 1024 > 100:
        #                         df2_ = pd.read_excel(file)
        #                         print('#1', file)
        #                         if len(df1_) != 0:
        #                             df2_ = df2_.iloc[10:]
        #                         if len(weekends) > 1:
        #                             df2_ = df2_.iloc[:-1]
        #                         df1_ = pd.concat([df1_, df2_])
        #                         break
        #             except:
        #
        #                 if day in subfolders and 'kzt народный' in subfolders.lower() and os.path.getsize(subfolder) / 1024 > 100:
        #                     df2_ = pd.read_excel(subfolder)
        #                     print('#2', subfolder)
        #                     if len(df1_) != 0:
        #                         df2_ = df2_.iloc[10:]
        #                     if len(weekends) > 1:
        #                         df2_ = df2_.iloc[:-1]
        #                     df1_ = pd.concat([df1_, df2_])
        #
        # print('LENTUS', len(df1_))

        for folder, subfolder, files in os.walk(halyk_extract_path):
            for file in files:
                # print(os.path.join(folder, file))
                if day in file and 'kzt народный' in file.lower() and os.path.getsize(os.path.join(folder, file)) / 1024 > 100:
                    df2_ = pd.read_excel(os.path.join(folder, file))
                    print('#1', file)
                    if len(df1_) != 0:
                        df2_ = df2_.iloc[10:]
                    if len(weekends) > 1:
                        df2_ = df2_.iloc[:-1]
                    df1_ = pd.concat([df1_, df2_])
                    break

        # * --- Remove below code ---
        if len(df1_) == 0:

            for folder, subfolder, files in os.walk(halyk_extract_path_2023):
                for file in files:
                    # print(os.path.join(folder, file))
                    if day in file and 'kzt народный' in file.lower() and os.path.getsize(os.path.join(folder, file)) / 1024 > 100:
                        df2_ = pd.read_excel(os.path.join(folder, file))
                        print('#1', file)
                        if len(df1_) != 0:
                            df2_ = df2_.iloc[10:]
                        if len(weekends) > 1:
                            df2_ = df2_.iloc[:-1]
                        df1_ = pd.concat([df1_, df2_])
                        break
        # * --- ---

    df1_.dropna(how='all', inplace=True)

    if True:

        df1_.columns = df1_.iloc[7]

        if len(weekends) > 1:
            df1_ = df1_[(df1_['Дебет'].notna()) | (df1_['Кредит'].notna())].iloc[1:]
            # df1_ = df1_[(df1_['Дебет'].notna())].iloc[1:]
        else:
            df1_ = df1_[(df1_['Дебет'].notna()) | (df1_['Кредит'].notna())].iloc[1:-1]
            # df1_ = df1_[(df1_['Дебет'].notna())].iloc[1:-1]

        debit = df1_[df1_['Дебет'].notna()].copy()
        credit = df1_[df1_['Кредит'].notna()].copy()

        with suppress(Exception):

            debit['Дебет'] = debit['Дебет'].apply(lambda x: x.replace(' ', ''))
            credit['Кредит'] = credit['Кредит'].apply(lambda x: x.replace(' ', ''))

        try:
            debit['Дебет'] = debit['Дебет'].astype(float)
            credit['Кредит'] = credit['Кредит'].astype(float)

            # df1_['Дебет'] = df1_['Дебет'].astype(float)
            # df1_['Кредит'] = df1_['Кредит'].astype(float)
        except Exception as er:
            print(er)

        # print('-------')
        # print(debit)
        # print('-===-')
        # print(credit)
        # print('--===--')
        # print(pd.concat([debit, credit]))
        print()

        df1_ = pd.concat([debit, credit])

        df1_['Дата валютирования'] = pd.to_datetime(df1_['Дата валютирования'], format='%d.%m.%Y')
        df1_['Дата валютирования'] = df1_['Дата валютирования'].dt.strftime('%d.%m.%y')

        # print(len(df1_))
        # df1_.to_excel('rffdgdlolus.xlsx')

        return df1_

    # except:
    #     send_message_to_orc('https://rpa.magnum.kz/tg', chat_id, 'Сверка выписок\nОШИБКА4')


def odines(yesterdays_reestr_date):
    # send_message_to_orc('https://rpa.magnum.kz/tg', chat_id, f'Начат 1С')
    app = Odines()

    app.auth()
    app.find_element({"title": "БДР", "class_name": "", "control_type": "Button", "visible_only": True,
                      "enabled_only": True, "found_index": 0}).click()
    time.sleep(0.1)
    app.find_element({"title": "Реестр платежей", "class_name": "", "control_type": "MenuItem", "visible_only": True,
                      "enabled_only": True, "found_index": 0}).click()

    app.wait_element({"title": "Установить период...", "class_name": "", "control_type": "Button", "visible_only": True,
                      "enabled_only": True, "found_index": 0}, timeout=2)

    app.find_element({"title": "Установить период...", "class_name": "", "control_type": "Button", "visible_only": True,
                      "enabled_only": True, "found_index": 0}).click()

    app.wait_element({"title": "", "class_name": "", "control_type": "Edit", "visible_only": True,
                      "enabled_only": True, "found_index": 0}, timeout=10)

    app.switch({"title": "Выберите период", "class_name": "V8NewLocalFrameBaseWnd", "control_type": "Window", "visible_only": True, "enabled_only": True, "found_index": 0})

    yesterdays_reestr_date_ = yesterdays_reestr_date.replace('.', '')[:4] + yesterdays_reestr_date.replace('.', '')[-2:]

    df = pd.DataFrame(columns=cols)

    for i in range(5):

        if True:

            app.find_element({"title": "", "class_name": "", "control_type": "Edit", "visible_only": True,
                              "enabled_only": True, "found_index": 0}).type_keys(yesterdays_reestr_date_, app.keys.TAB)

            app.find_element({"title": "", "class_name": "", "control_type": "Edit", "visible_only": True,
                              "enabled_only": True, "found_index": 1}).type_keys(yesterdays_reestr_date_, app.keys.TAB)

            app.find_element({"title": "Выбрать", "class_name": "", "control_type": "Button",
                              "visible_only": True, "enabled_only": True, "found_index": 0}, timeout=15).click()

            app.switch({"title": "", "class_name": "", "control_type": "Pane", "visible_only": True, "enabled_only": True, "found_index": 29})

            if not app.wait_element({"title_re": ".*Дата", "class_name": "", "control_type": "Custom", "visible_only": True, "enabled_only": True, "found_index": 0}, timeout=15):
                print('quitting')
                # time.sleep(1000)
                app.quit()

                return pd.DataFrame()

            print('found')
            keyboard.press_and_release('ctrl+a')

            time.sleep(.2)

            tools.clipboard_set('')
            while len(pyperclip.paste()) == 0:
                keyboard.press_and_release('ctrl+c')
                time.sleep(0.3)

            all_reestrs_in_1c = csv.reader(pyperclip.paste().splitlines(), delimiter='\t')  # считывание таблицы из 1С в пандас
            dt = [row1 for row1 in all_reestrs_in_1c]

            all_reestrs_in_1c = pd.DataFrame(dt)

            print(all_reestrs_in_1c, len(all_reestrs_in_1c))

            keyboard.press_and_release('page_down')

            for i in range(len(all_reestrs_in_1c)):

                keyboard.press_and_release('enter')
                app.switch({"title": "", "class_name": "", "control_type": "Pane", "visible_only": True,
                            "enabled_only": True, "found_index": 36})

                keyboard.press_and_release('tab')
                time.sleep(.1)
                keyboard.press_and_release('tab')
                time.sleep(.1)
                keyboard.press_and_release('tab')
                time.sleep(.1)
                keyboard.press_and_release('tab')
                time.sleep(.1)

                keyboard.press_and_release('ctrl+a')

                time.sleep(.2)

                tools.clipboard_set('')
                while len(pyperclip.paste()) < 10:
                    keyboard.press_and_release('ctrl+c')
                    time.sleep(0.3)

                df1 = csv.reader(pyperclip.paste().splitlines(), delimiter='\t')  # считывание таблицы из 1С в пандас
                dt = [row1 for row1 in df1]

                df1 = pd.DataFrame(dt)
                df1.columns = cols

                tools.clipboard_set('')
                # row = df1.columns
                # df1.columns = cols
                # df1.loc[0:0] = row
                df = pd.concat([df, df1], ignore_index=True)

                keyboard.press_and_release('esc')

                time.sleep(0.2)

                keyboard.press_and_release('up')

                time.sleep(0.1)

            break

        # except:
        #     pass
    # При копировании из 1С столбец Согласован пропадает и столбцы дат (не понял почему только они) смещаются на 1 вправо!!!
    df['Дата выписки'] = df['Дата выписки'].apply(lambda x: x.rstrip('.1'))
    df['Дата выписки'] = df['Дата выписки'].apply(lambda x: x[:6] + x[-2:])

    # * Поптыка маппинга с 1С и листом Статьи
    # df_temp = pd.read_excel(f'{save_xlsx_path}\\Шаблоны для робота (не удалять)\\Копия Сверка ОБРАЗЕЦ.xlsx', sheet_name=2)
    #
    # df_temp.columns = ['', 'Статья в ДДС', 'Код', 'Название статьи'] + [''] * (len(df_temp.columns) - 4)
    #
    # df_temp = df_temp.dropna(subset=['Статья в ДДС'])
    # df_temp.to_excel('chpokus1.xlsx')
    # df2_ = df.copy()
    # df2_.to_excel('chpokus2.xlsx')
    # df2_ = df2_.merge(df_temp, left_on='Код БДДС', right_on='Код', how='left')
    # print(df2_)
    # df2_.to_excel('chpokus.xlsx')

    df['Сумма документа'] = df['Сумма документа'].apply(lambda x: re.sub(r'\s+', '', x.replace(',', '.')))

    df1_ = pd.DataFrame({'Поставщик': df['Контрагент'], 'БИН / ИИН получателя': df['БИН / ИИН'].astype(str), 'Реестр': 'Реестр 1С', 'Статья в ДДС': df['Статья затрат'], 'Валюта платежа': 'KZT',
                         'Сумма к оплате': df['Сумма документа'].astype(float), 'Сумма к оплате KZT': df['Сумма документа'].astype(float), 'Курс': 1, 'Дата оплаты': df['Дата выписки'], 'Skip': '', 'Проверка статьи': df['Код БДДС'], 'Название статьи': df['Статья затрат']})
    app.quit()
    time.sleep(1)

    return df1_


def design_number_fmt_and_date(df2, yest):
    logger.info('Started designing number and date formats')
    # print('Started designing number and date formats')
    # send_message_to_orc('https://rpa.magnum.kz/tg', chat_id, f'Начато форматирование ячеек для чисел и дат')

    book = load_workbook(f'{working_path}\\Temp1.xlsx')  # edit1

    book.active = book['Реестры']
    sheet = book.active

    rows = df2.to_numpy().tolist()

    for r_idx, row in enumerate(rows, 19):
        for c_idx, value in enumerate(row, 1):
            sheet[f'B{r_idx}'].number_format = '0'
            sheet.cell(row=r_idx, column=c_idx, value=str(value)).number_format = '0'

    sheet['D1'] = yest

    book.save(f'{working_path}\\Temp1.xlsx')  # edit1
    book.close()


def fill_empty_bins():
    # print('Started filling empty bins')
    # send_message_to_orc('https://rpa.magnum.kz/tg', chat_id, f'Начато заполнение пустых БИНов')
    book = load_workbook(f'{working_path}\\Temp1.xlsx')

    sheet = book['БИНы и Компании']

    bins = []
    companies = []

    for ind in range(2, sheet.max_row):
        if sheet[f'A{ind}'].value is None and sheet[f'B{ind}'].value is None:
            break
        bins.append(sheet[f'A{ind}'].value)
        companies.append(sheet[f'B{ind}'].value)

    sheet = book['Реестры']
    for i in range(19, sheet.max_row):
        if sheet[f'A{i}'].value is None and sheet[f'B{i}'].value is None:
            break

        for ind, company in enumerate(companies):
            # if sheet[f'A{i}'].value == 'Научно-производственное Объединение Дортехника ТОО' == company:
            #     print(company, sheet[f'B{i}'].value, bins[ind])
            if company == sheet[f'A{i}'].value and sheet[f'B{i}'].value is None:
                sheet[f'B{i}'].value = bins[ind]

    book.save(f'{working_path}\\Temp1.xlsx')
    book.close()

    time.sleep(0.3)


def make_analysis_and_calculations(yesterday):
    hold_session()
    logger.info('Started making analysis and calculations')
    # print('Started making analysis and calculations')
    # send_message_to_orc('https://rpa.magnum.kz/tg', chat_id, f'Начаты анализ и подсчёт файла')
    if True:  # Temp2323
        book = load_workbook(f'{working_path}\\Temp1.xlsx')

        ids = []
        # contragent_name_halyk = []
        # payment_amount_halyk = []
        # payment_purpose_halyk = []
        # bin_halyk = []

        max_rows_halyk = 0

        sheet = book['Halyk']

        for i in range(2, sheet.max_row):

            if sheet[f'E{str(i)}'].value is None or sheet[f'E{str(i)}'].value == '':
                max_rows_halyk = i
                break

            if sheet[f'H{str(i)}'].value is not None:
                ids.append(i)
                # contragent_name_halyk.append(sheet[f'E{str(i)}'].value)
                # payment_amount_halyk.append(float(sheet[f'H{str(i)}'].value))
                # payment_purpose_halyk.append(str(sheet[f'J{str(i)}'].value))
                # bin_halyk.append(sheet[f'F{str(i)}'].value)
        # for _ in ids:
        #     print(_)
        matches = []

        sheet = book['Реестры']
        sheet_halyk = book['Halyk']

        hold_session()
        for i in range(19, sheet.max_row):
            if sheet[f'A{i}'].value is None:
                break

            contragent_name_reestr = sheet[f'A{i}'].value
            payment_amount_reestr = sheet[f'G{i}'].value
            currency = sheet[f'E{i}'].value
            reestr_bin = sheet[f'B{i}'].value

            try:
                payment_amount_reestr = payment_amount_reestr.replace(' ', '')
                reestr_bin = reestr_bin.strip()
            except:
                ...

            sheet[f'Q{i}'].value = 'Не идёт'

            for ind in ids:

                try:
                    bin_halyk = str(sheet_halyk[f'F{ind}'].value)
                    bin_halyk = bin_halyk.strip()
                    bin_halyk = bin_halyk.lstrip('0')
                except:
                    ...
                try:
                    reestr_bin = str(reestr_bin)
                    reestr_bin = reestr_bin.lstrip('0')
                except:
                    ...
                try:
                    payment_amount_reestr = float(payment_amount_reestr)
                    payment_amount_reestr = round(payment_amount_reestr, 3)
                except:
                    ...
                payment_amount_halyk = round(float(sheet_halyk[f'H{ind}'].value), 3)

                contragent_name_halyk = sheet_halyk[f'E{ind}'].value
                payment_purpose_halyk = str(sheet_halyk[f'J{ind}'].value)

                if 'комиссия' in payment_purpose_halyk.lower():
                    matches.append(ind)
                    continue

                # 2 пункт
                if ('покупка' in payment_purpose_halyk.lower() and 'валют' in payment_purpose_halyk.lower()) and currency in ['USD', 'EUR', 'RUB'] and contragent_name_halyk == contragent_name_reestr and payment_amount_halyk == payment_amount_reestr:
                    sheet[f'Q{str(i)}'].value = 'Да'
                    matches.append(ind)
                    continue

                elif contragent_name_halyk == contragent_name_reestr and payment_amount_halyk == payment_amount_reestr:
                    sheet[f'Q{str(i)}'].value = 'Да'
                    matches.append(ind)
                    continue

                # 3.1
                elif ('ао' in contragent_name_halyk.lower() and 'народный' in contragent_name_halyk.lower() and 'банк' in contragent_name_halyk.lower() and 'казахстана' in contragent_name_halyk.lower()) and \
                        contragent_name_reestr.lower() == 'сотрудники' and payment_amount_halyk == payment_amount_reestr:
                    sheet[f'Q{str(i)}'].value = 'Да'
                    # print('---', payment_amount_reestr, payment_amount_halyk[ind])
                    matches.append(ind)
                    continue

                if ('погашение со счета' in payment_purpose_halyk.lower() or 'проценты по кредиту' in payment_purpose_halyk.lower() or 'выдача размена' in payment_purpose_halyk.lower() or 'для зачисления на картсчета сотрудникам' in payment_purpose_halyk.lower()) \
                        and payment_amount_halyk == payment_amount_reestr and (bin_halyk == reestr_bin or contragent_name_halyk == contragent_name_reestr):
                    matches.append(ind)
                    continue

                # 4
                elif (bin_halyk == reestr_bin) and payment_amount_halyk == payment_amount_reestr:
                    sheet[f'Q{str(i)}'].value = 'Да'
                    matches.append(ind)
                    continue

                # ПРОВЕРКА НА СХОЖЕСТЬ СТРОК
                elif contragent_name_halyk in contragent_name_reestr and payment_amount_halyk == payment_amount_reestr:
                    sheet[f'Q{str(i)}'].value = 'Да'
                    matches.append(ind)
                    continue

                else:
                    match = 0

                    for row1 in contragent_name_halyk.split():
                        for symbol in ",'!@#$%^&*()_+-=-./?|<>[]{}:;\"":
                            row1 = row1.replace(symbol, ' ')

                        for row2 in contragent_name_reestr.split():
                            for symbol in ",'!@#$%^&*()_+-=-./?|<>[]{}:;№«»\"":
                                row2 = row2.replace(symbol, ' ')

                            if row1.lower().strip() == row2.lower().strip():
                                match += 1

                    num = max(len(contragent_name_halyk.split()), len(contragent_name_reestr.split()))
                    if match * 100.0 / num >= 100 and payment_amount_halyk == payment_amount_reestr:
                        # print('СХОЖИ: ', contragent_name_halyk[ind], ' | ', contragent_name_reestr)
                        sheet[f'Q{str(i)}'].value = 'Да'
                        matches.append(ind)
                        continue

        hold_session()

        for i in range(19, sheet.max_row):
            if sheet[f'F{i}'].value is None and sheet[f'G{i}'].value is None:
                break
            try:
                sheet[f'F{i}'].number_format = 'General'
                sheet[f'F{i}'].value = float(sheet[f'F{i}'].value)
                sheet[f'F{i}'].number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
            except:
                ...

            try:
                sheet[f'G{i}'].number_format = 'General'
                sheet[f'G{i}'].value = float(sheet[f'G{i}'].value)
                sheet[f'G{i}'].number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
            except:
                ...

        sheet = book['Halyk']

        not_matching = np.arange(2, max_rows_halyk)
        matches = np.unique(np.array(matches))
        # print(max_rows_halyk)
        # for _ in matches:
        #     print(_)

        for ind in not_matching:
            if sheet[f'H{ind}'].value is not None:
                sheet[f'O{ind}'].value = str('Не идёт')

        for ind in matches:
            # if sheet[f'H{ind + 2}'].value is not None:
            sheet[f'O{ind}'].value = str('Да')

        for ind in range(2, sheet.max_row):
            if sheet[f'A{ind}'] is None and sheet[f'H{ind}'] is None:
                break
            sheet[f'H{ind}'].number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

        book.save(f'{working_path}\\Temp1.xlsx')
        book.close()
        time.sleep(1)
        try:
            os.system('taskkill /im excel.exe /f')
        except:
            ...
        time.sleep(1)

        excel_app = xw.App(visible=False)
        book = excel_app.books.open(f'{working_path}\\Temp1.xlsx', corrupt_load=True)

        app = xw.apps.active
        app.calculate()

        sheet = book.sheets['Реестры']
        # print('Started clearing Реестры')

        rng = sheet.range('A19')
        max_row = max(rng.current_region.end('down').row, rng.end('down').row)
        ind = max_row

        cell = f'A{20}:L{ind}'
        sheet.range(cell).font.name = 'Calibri'
        sheet.range(cell).font.size = '11'
        # print('LEN: ', ind)

        hold_session()

        cell = f'N{10001}:W{10001}'
        print(cell)

        for tries in range(30):
            try:
                sheet.range(cell).clear_contents()
                sheet.range(cell).clear_formats()
                break
            except:
                time.sleep(13)

        # print('Started clearing Halyk')
        sheet = book.sheets['Halyk']

        rng = sheet.range('A2')
        max_row = max(rng.current_region.end('down').row, rng.end('down').row)
        ind1 = max_row

        cell = f'K{ind1 + 1}:W{10001}'
        sheet.range(cell).clear_contents()
        sheet.range(cell).clear_formats()
        # send_message_to_orc('https://rpa.magnum.kz/tg', chat_id, f'Всё сверено. Лишние строки были удалены: длина {ind} - Реестры, {ind1} - Halyk')
        sheet.range(f'L1:O{ind1 + 1}').api.VerticalAlignment = VAlign.xlVAlignCenter
        sheet.range(f'L1:O{ind1 + 1}').api.HorizontalAlignment = HAlign.xlHAlignCenter

        folder = None
        month = int(yesterday.split('.')[1]) - 1
        year = yesterday.split('.')[2]

        with suppress(Exception):
            f_path = os.path.join(os.path.join(save_xlsx_path, 'сверка 2023'), f'{MONTHS[month]} {year}') if int(year) == 2023 else os.path.join(save_xlsx_path, f'{MONTHS[month]} {year}')
            print(f_path)
            os.makedirs(f_path, exist_ok=True)
            folder = f_path
        print(folder)
        for tries in range(5):
            try:
                book.save(f'{folder}\\Сверка {yesterday}.xlsx')
                book.save(f'{save_xlsx_path_qlik}\\Сверка {yesterday}.xlsx')
                break
            except:
                time.sleep(15)

        try:
            book.close()
            app.quit()
            app.kill()
        except:
            ...
        try:
            os.remove(f'{working_path}\\Temp1.xlsx')
        except:
            ...

        return [ind, ind1]

    # except:
    #     send_message_to_orc('https://rpa.magnum.kz/tg', chat_id, 'Сверка выписок\nОШИБКА5')


# ORIGIN CODE
if __name__ == '__main__':

    logger = logging.getLogger(logger_name)

    try:
        start_time = datetime.datetime.now().strftime('%H:%M:%S')
        start_time_secs = time.time()
        timings = []
        start_time_iter = datetime.datetime.now().strftime('%H:%M:%S')

        update_credentials(save_xlsx_path, owa_username, owa_password)
        update_credentials(save_xlsx_path_qlik, owa_username, owa_password)

        for month_ in range(1):
            day__ = [3, 10]
            for day in range(1):

                yesterday1 = datetime.date.today().strftime('%d.%m.%y')
                yesterday2 = datetime.date.today().strftime('%d.%m.%Y')

                if process_day != '':
                    yesterday1 = process_day
                    yesterday2 = f"{process_day.split('.')[0]}.{process_day.split('.')[1]}.{process_day.split('.')[2][2:]}"

                # Если вдруг надо будет запускать на большой период дат
                # if day < 10:
                #     yesterday1 = f'0{day}.02.24'
                #     yesterday2 = f'0{day}.02.2024'
                # else:
                #     yesterday1 = f'{day}.02.24'
                #     yesterday2 = f'{day}.02.2024'

                logger.info(f'STARTED DAY: {yesterday1}, {yesterday2}')

                calendar = pd.read_excel(f'{save_xlsx_path}\\Шаблоны для робота (не удалять)\\Производственный календарь {yesterday2[-4:]}.xlsx')

                cur_day_index = calendar[calendar['Day'] == yesterday1]['Type'].index[0]
                cur_day_type = calendar[calendar['Day'] == yesterday1]['Type'].iloc[0]

                if cur_day_type != 'Holiday':
                    # print('Started current date: ', yesterday2)
                    weekends = []
                    weekends_type = []

                    for i in range(cur_day_index - 1, -1, -1):
                        print(i, calendar['Day'].iloc[i][:6] + '20' + calendar['Day'].iloc[i][-2:])
                        weekends.append(calendar['Day'].iloc[i][:6] + '20' + calendar['Day'].iloc[i][-2:])
                        weekends_type.append(calendar['Type'].iloc[i])
                        if calendar['Type'].iloc[i] == 'Working':
                            yesterday1 = calendar['Day'].iloc[i]
                            break

                    df = get_first_statement(weekends)

                    book = load_workbook(f'{save_xlsx_path}\\Шаблоны для робота (не удалять)\\Копия Сверка ОБРАЗЕЦ.xlsx')

                    book.active = book['Halyk']
                    sheet = book.active

                    rows = df.to_numpy().tolist()

                    for r_idx, row in enumerate(rows, 2):
                        for c_idx, value in enumerate(row, 1):
                            sheet.cell(row=r_idx, column=c_idx, value=value)

                    book.save(f'{working_path}\\Temp1.xlsx')

                    df3 = pd.DataFrame()

                    for ind, yesterday in enumerate(weekends):
                        # # 1 --------------------------------------------------------------------------
                        print('yes:', yesterday)
                        web1 = search_by_date(yesterday)

                        # # 2 --------------------------------------------------------------------------
                        # df2 = pd.DataFrame()
                        # yesterdays_reestr_date = '30.10.2023'
                        print('finished filter')
                        df2, yesterdays_reestr_date = documentolog(web1, yesterday)
                        print('finished sed')
                        # # 3 --------------------------------------------------------------------------

                        isEmpty = False

                        if weekends_type[ind] != 'Holiday' and df2 is not None and yesterdays_reestr_date is not None:

                            df1 = odines(yesterday)

                            df2 = pd.concat([df2, df1])

                            if len(df1) == 0:
                                isEmpty = True
                                tools.send_message_to_tg(tg_token, chat_id, f'Реестры 1С - Пустые')

                        df3 = pd.concat([df3, df2])

                    # 4 ---------------------------------------------------------------------------------------

                    design_number_fmt_and_date(df3, yesterday1)

                    # 5 ---------------------------------------------------------------------------------------

                    fill_empty_bins()

                    # 6 ---------------------------------------------------------------------------------------

                    len_reestr, len_halyk = make_analysis_and_calculations(yesterday2)

                    # # FINISHED LOGIC --------------------------------------------------------------------------

                    tools.send_message_to_tg(tg_token, chat_id, f'Всё сверено. Отрабатывал за сегодня({yesterday2}), день(дни) за которые брал реестры {weekends}\nЛишние строки были удалены\nОбщая длина Реестров - {len_reestr}, Halyk - {len_halyk}')

                    # send_message_by_smtp(smtp_host, to=['Abdykarim.D@magnum.kz', 'Goremykin@magnum.kz', 'Ibragimova@magnum.kz'], subject=f'Сверка Выписок ROBOT - {yesterday2}', body=f'Сверка Выписок за {yesterday2} завершилась', username=smtp_author)

                else:
                    print(1)

    except Exception as main_error:
        print(f'MAIN ERROR OCCURED: {main_error}')
        logger.info(f'MAIN ERROR OCCURED: {main_error}')
        logger.warning(f'MAIN ERROR OCCURED: {main_error}')
