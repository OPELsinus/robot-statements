import ctypes
import inspect
import io
import json
import logging
import os
import random
import re
import shutil
import smtplib
import socket
import string as string_lib
import subprocess
import time
import traceback
import urllib.parse
from contextlib import suppress
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from zipfile import ZipFile

import pyautogui
import requests
import telegram
import win32clipboard
from openpyxl import load_workbook
from openpyxl.styles import Side, Border, Font
from pywinauto.timings import wait_until_passes
from win32gui import GetCursorPos

NUMBER_FORMAT = '# ##0.00_-'


def remove_credentials():
    command = f'net use * /delete /y'
    result = subprocess.run(command, shell=True, capture_output=True, encoding='cp866')
    if result.returncode != 0:
        logging.warning(str(result.stderr).replace('\n', ' '))
    else:
        logging.info(str(result.stdout).replace('\n', ' '))


def update_credentials(resource: Path, username, password, clear=False):
    if clear:
        remove_credentials()
    resource = str(resource)[:-1] if str(resource)[-1] == '\\' else str(resource)
    command = f'net use "{resource}" /user:{username} {password}'
    result = subprocess.run(command, shell=True, capture_output=True, encoding='cp866')
    if result.returncode != 0:
        logging.warning(str(result.stderr).replace('\n', ' '))
    else:
        logging.info(str(result.stdout).replace('\n', ' '))


def json_read(path: Path):
    with open(str(path), 'r', encoding='utf-8') as fp:
        data = json.load(fp)
    return data


def json_write(path: Path, data):
    with open(str(path), 'w', encoding='utf-8') as fp:
        json.dump(data, fp, ensure_ascii=False)


def get_hostname():
    hostname = socket.gethostbyname(socket.gethostname())
    return hostname


def check_session():
    try:
        wait_until_passes(0.1, 0.1, GetCursorPos)
        return True
    except (Exception,):
        return False


def prevent_auto_lock():
    ctypes.windll.kernel32.SetThreadExecutionState(0x80000000 | 0x00000001)


def send_message_to_tg(bot_token, chat_id, message):
    import requests

    r = requests.post(f"https://api.telegram.org/bot{bot_token}/sendMessage", json={'chat_id': chat_id, 'text': message}, verify=False)

def take_screenshot():
    screenshot = pyautogui.screenshot()
    scr_bytes = io.BytesIO()
    screenshot.save(scr_bytes, format='PNG')
    scr_bytes.seek(0)

    return scr_bytes

def send_screenshot_to_tg(bot_token, chat_id, screenshot_bytes):
    import requests

    response = requests.post(f"https://api.telegram.org/bot{bot_token}/sendPhoto", data={'chat_id': chat_id}, files={'photo': ('screenshot.png', screenshot_bytes)}, verify=False)


def send_message_to_orc(url, chat_id, message, log=True):
    if log:
        requests.post(url, data={'chat_id': chat_id, 'message': message}, verify=False)


def protect_string_for_path(string):
    return re.sub(r'[<>:"/\\|?*]', '_', string)


def protect_string_for_url(url):
    return urllib.parse.quote(url, safe='/:')


def remove_multiple_newlines_spaces(string):
    return re.sub(r'\n{2,}|\s{2,}', ' ', str(string))


def generate_random_filename(length, extension=None):
    random_string = ''.join(random.choice(string_lib.ascii_letters + string_lib.digits) for _ in range(length))
    if extension:
        return f'{random_string}.{extension}'
    else:
        return random_string


def check_file_downloaded(folder_path, file_name, check_interval=1, timeout=60):
    start_time = time.time()
    while True:
        folder = Path(folder_path)
        files = folder.glob(file_name)
        for file_path in files:
            if not any(temp in str(file_path) for temp in ['.crdownload', '~$']):
                if file_path.is_file() and file_path.stat().st_size > 0:
                    return file_path
        if time.time() - start_time > timeout:
            return None
        time.sleep(check_interval)
        with suppress(Exception):
            pyautogui.press('volumedown')
            time.sleep(0.1)
            pyautogui.press('volumeup')


def fix_excel_file_error(file_path: Path):
    try:
        tmp_folder = file_path.parent.joinpath('__temp__')
        with ZipFile(file_path.__str__()) as excel_container:
            excel_container.extractall(tmp_folder)
            excel_container.close()
        wrong_file_path = os.path.join(tmp_folder.__str__(), 'xl', 'SharedStrings.xml')
        correct_file_path = os.path.join(tmp_folder.__str__(), 'xl', 'sharedStrings.xml')
        os.rename(wrong_file_path, correct_file_path)
        file_path.unlink()
        shutil.make_archive(file_path.__str__(), 'zip', tmp_folder)
        os.rename(file_path.__str__() + '.zip', file_path.__str__())
        shutil.rmtree(tmp_folder.__str__(), ignore_errors=True)
    except Exception as e:
        traceback.print_exc()
        logging.warning(f"Error while trying to fix excel file: {e}")
        return None
    with suppress(Exception):
        pyautogui.press('volumedown')
        time.sleep(0.1)
        pyautogui.press('volumeup')
    return file_path


def send_message_by_smtp(smtp_host, to, subject, body, username, password=None, html=None, attachments=None):
    with smtplib.SMTP(smtp_host, 25) as smtp:
        smtp.ehlo()
        smtp.starttls()
        smtp.ehlo()
        if password:
            smtp.login(username, password)

        msg = MIMEMultipart('alternative')
        msg["From"] = username
        msg["To"] = ';'.join(to) if type(to) is list else to
        msg["Subject"] = subject
        msg.attach(MIMEText(body, 'plain'))

        if html:
            msg.attach(MIMEText(html, 'html'))

        if attachments and isinstance(attachments, list):
            for each in attachments:
                path = Path(each).resolve()
                with open(path.__str__(), 'rb') as f:
                    part = MIMEApplication(f.read(), Name=path.name)
                    part['Content-Disposition'] = 'attachment; filename="%s"' % path.name
                    msg.attach(part)

        smtp.send_message(msg=msg)


def get_parent_function_info():
    stack = inspect.stack()
    parent_frame = stack[1]
    return parent_frame.function, parent_frame.lineno


def xlsx_read(path: Path, sheet_name=None, data_only=True):
    wb = load_workbook(path.__str__(), data_only=data_only)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            return list()
        else:
            ws = wb[sheet_name]
    else:
        ws = wb.active
    return list(ws.values)


def xlsx_find(path: Path, value, sheet_name=None, row_index=None, col_index=None):
    wb = load_workbook(path.__str__(), data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            return list()
        else:
            ws = wb[sheet_name]
    else:
        ws = wb.active
    rows = [row for row in list(ws.values)]
    pairs = list()
    for n, row in enumerate(rows):
        col_indexes = [(n + 1, i + 1) for i, x in enumerate(row) if x == value]
        pairs = [*pairs, *col_indexes]
    if row_index:
        pairs = [pair for pair in pairs if pair[0] == row_index]
    if col_index:
        pairs = [pair for pair in pairs if pair[1] == col_index]
    return pairs


def xlsx_get(path: Path, row, col, sheet_name=None):
    wb = load_workbook(path.__str__(), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    return ws.cell(row=row, column=col).value


def xlsx_set(path: Path, row, col, value, sheet_name=None, timeout=60, number_format=None, border=True, font=None):
    def function():
        wb = load_workbook(path.__str__(), data_only=True)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet()
                ws.title = sheet_name
            else:
                ws = wb[sheet_name]
        else:
            ws = wb.active
        if number_format:
            ws.cell(row=row, column=col).number_format = number_format
        if border:
            thin = Side(style='thin')
            ws.cell(row=row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.cell(row=row, column=col).font = font or Font(name='Calibri', size=11)
        ws.cell(row=row, column=col).value = value

        temp_path = path.parent.joinpath(f'~${path.name}')

        start_time = time.time()
        while True:
            if not temp_path.is_file():
                break
            if time.time() - start_time > timeout:
                raise Exception(f'{path} is read only')
            time.sleep(1)
        temp_path = path.parent.joinpath(f'_{path.name}')
        shutil.copy(path.__str__(), temp_path)
        wb.save(path.__str__())
        wb.close()
        temp_path.unlink()
    wait_until_passes(60, 1, function)


def xlsx_append(path: Path, row, sheet_name=None, timeout=60):
    def function():
        wb = load_workbook(path.__str__(), data_only=True)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet()
                ws.title = sheet_name
            else:
                ws = wb[sheet_name]
        else:
            ws = wb.active
        ws.append(row)
        temp_path = path.parent.joinpath(f'~${path.name}')

        start_time = time.time()
        while True:
            if not temp_path.is_file():
                break
            if time.time() - start_time > timeout:
                raise Exception(f'{path} is read only')
            time.sleep(1)
        temp_path = path.parent.joinpath(f'_{path.name}')
        shutil.copy(path.__str__(), temp_path)
        wb.save(path.__str__())
        wb.close()
        temp_path.unlink()
    wait_until_passes(60, 1, function)


def clipboard_get(raise_err=True):
    win32clipboard.OpenClipboard()
    result = win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
    win32clipboard.CloseClipboard()
    if not len(result):
        if raise_err:
            raise Exception('Clipboard is empty')
        else:
            return None
    return result


def clipboard_set(value):
    win32clipboard.OpenClipboard()
    win32clipboard.EmptyClipboard()
    win32clipboard.SetClipboardData(win32clipboard.CF_UNICODETEXT, value)
    win32clipboard.CloseClipboard()


def remove_chars(value, deletechars=None):
    deletechars = deletechars or r'\/:*?"<>|'
    for c in deletechars:
        value = value.replace(c, '')
    return value


def hold_session():
    with suppress(Exception):
        pyautogui.press('volumedown')
        pyautogui.press('volumeup')


class PostHandler(logging.Handler):
    def __init__(self, url, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.url = url

    def emit(self, record):
        data = self.format(record)
        requests.post(self.url, json=data, verify=False)
